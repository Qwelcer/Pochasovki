import openpyxl
import sqlite3

wb = openpyxl.open("почасовые отчёты.xlsx")
print(wb.active)
sheet = wb.active
sheet_max_row = sheet.max_row

try:
    sqlite_connection = sqlite3.connect('index.db')
    cursor = sqlite_connection.cursor()
    print("Connection success")

    record = cursor.fetchall()
    print("Версия базы данных SQLite: ", record)

    for i in range(1, sheet_max_row):
        consumer = sheet.cell(row=i, column=1)
        consumer_value = consumer.value

        BS = sheet.cell(row=i, column=3)
        BS_value = BS.value

        address = sheet.cell(row=i, column=4)
        address_value = address.value

        measurer = sheet.cell(row=i, column=5)
        measurer_value = measurer.value

        last = sheet.cell(row=i, column=7)
        last_value = last.value

        consumption = sheet.cell(row=i, column=9)
        consumption_value = consumption.value

        insert = (consumer_value, BS_value, address_value, measurer_value, last_value, consumption_value)

        cursor.execute(
            "INSERT INTO pochasovki (consumer, bs, address, measurer, last, consumption) VALUES(?, ?, ?, ?, ?, ?);",
            insert)

    # cursor.execute("INSERT INTO pochasovki (potrebitel, bs, address, pu, feb22, rashod) VALUES(?, ?, ?, ?, ?, ?);", insert)
    sqlite_connection.commit()
    cursor.close()

except sqlite3.Error as error:
    print("Ошибка при подключении к sqlite", error)
finally:
    if sqlite_connection:
        sqlite_connection.close()
        print("Соединение с SQLite закрыто")
