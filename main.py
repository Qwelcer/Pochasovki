import openpyxl
import sqlite3
import random

wb = openpyxl.open("Шаблон 30.xlsx")
report = openpyxl.open("data report.xlsx")
sheet = wb.active
report_sheet = report.active

report_day = 'Сентябрь 2022'

try:
    sqlite_connection = sqlite3.connect('index.db')
    cursor = sqlite_connection.cursor()
    print("База данных успешно подключена к SQLite")

    cursor.execute("SELECT * FROM pochasovki")
    db_row = cursor.fetchall()
    row_count = len(db_row)

    for i in range(1, row_count + 1):
        draft = cursor.execute(
            f"SELECT consumer, address, measurer, last, consumption, bs FROM `pochasovki` WHERE `id` = '{i}'").fetchone()
        print(draft[5])

        consumption = draft[4]

        sheet['C2'] = draft[0]  # потребитель
        sheet['C3'] = draft[1]  # адрес
        sheet['C5'] = draft[1]  # адрес
        sheet['C6'] = "'" + str(draft[2])  # Прибор учета
        sheet['B12'] = draft[3]  # Последние показания
        #   sheet['D12'] = draft[3] + draft[4]  #   Точные конечные показания. Начальные показания + рандомизированный расход
        #   sheet['F12'] = draft[4]             #   Точный расход. Сейчас рандомизированный

        sheet['P13'] = report_day

        sheet_name_full = str(draft[0])

        sheet_name = sheet_name_full[8:] + '_30д'
        sheet.title = sheet_name


        report_day = 'Сентябрь 2022'

        consumption_mid = consumption / 720  # кол-во тиков в расчетном месяце (дни*24)
        consumption_day_mid = consumption / 30  # кол-во дней в расчитываемом месяце
        consumption_result = 0

        for i in range(2, 32):  # кол-во дней в месяце+2. первое число не трогать
            consumption_day = 0
            for j in range(16, 39):
                consumption_moment = random.uniform(consumption_mid * 0.9, consumption_mid)  # 0,9 отклонение от среднего
                sheet.cell(row=j, column=i).value = consumption_moment
                consumption_day = consumption_day + consumption_moment
                consumption_result = consumption_result + consumption_moment

            sheet.cell(row=39, column=i).value = consumption_day_mid - consumption_day  # последний компенсаторный час
            consumption_result = consumption_result + (consumption_day_mid - consumption_day)
            print(consumption_day, "   ", consumption_day_mid - consumption_day, "   ", consumption_day_mid, "    ",
                  consumption_result)

        #sheet['F12'] = round(consumption_result, 2)
        #sheet['AF40'] = round(consumption_result, 2)
        sheet['D12'] = round(draft[3] + consumption_result, 2)

        consumption_delta = consumption_result - consumption
        print(consumption_delta)

        file_name = sheet_name + ' ' + draft[1]
        wb.save(f"{file_name}.xlsx")

        # Report add part

        report_sheet.cell(row=i + 1, column=1).value = draft[0]  # потребитель ДЭС
        report_sheet.cell(row=i + 1, column=2).value = draft[5]  # БС
        report_sheet.cell(row=i + 1, column=3).value = draft[1]  # адрес
        report_sheet.cell(row=i + 1, column=4).value = draft[2]  # ПУ
        report_sheet.cell(row=i + 1, column=5).value = draft[3]  # Предыдущие показания
        report_sheet.cell(row=i + 1, column=6).value = consumption_result + draft[3]  # Последние показания
        report_sheet.cell(row=i + 1, column=7).value = consumption_result  # Расход
        report_sheet.cell(row=i + 1, column=8).value = consumption_delta  # Дельта

        print(consumption_mid)
        #break  # стереть эту строку, чтобы сделать все файлы

    # Доделать отчёт и как-нибудь привязать к ДБ

    # sqlite_connection.commit()
    cursor.close()
except sqlite3.Error as error:
    print("Ошибка при подключении к sqlite", error)
finally:
    if sqlite_connection:
        sqlite_connection.close()
        print("Соединение с SQLite закрыто")

report.save(f"'{report_day}'.xlsx")
